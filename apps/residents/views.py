# Residents views
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.core.mixins import NonBootstrapRequiredMixin
from apps.core.decorators import non_bootstrap_required
from django.db.models import Q
from django.http import HttpResponse
from .models import Resident
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date
from django.urls import reverse
from django.contrib import messages
from .forms import ResidentForm


class ResidentsListView(LoginRequiredMixin, NonBootstrapRequiredMixin, ListView):
    """
    Residents list view with search, filtering, and pagination.
    """
    model = Resident
    template_name = 'pages/residents/list.html'
    context_object_name = 'residents'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Resident.objects.filter(is_active=True).select_related('household_head')
        
        # Search functionality
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(middle_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(address__icontains=search_query) |
                Q(mobile_number__icontains=search_query)
            )
        
        # Filter by purok
        purok_filter = self.request.GET.get('purok', '').strip()
        if purok_filter and purok_filter != 'all':
            queryset = queryset.filter(purok=purok_filter)
        
        # Filter by sector
        sector_filter = self.request.GET.get('sector', '').strip()
        if sector_filter and sector_filter != 'all':
            sector_mapping = {
                'senior': Q(is_senior_citizen=True),
                'pwd': Q(is_pwd=True),
                'solo_parent': Q(is_solo_parent=True),
                '4ps': Q(is_4ps=True),
                'indigent': Q(is_indigent=True),
                'voter': Q(is_voter=True),
            }
            if sector_filter in sector_mapping:
                queryset = queryset.filter(sector_mapping[sector_filter])
        
        # Filter by age range
        age_min = self.request.GET.get('age_min', '').strip()
        age_max = self.request.GET.get('age_max', '').strip()
        
        if age_min or age_max:
            today = date.today()
            
            if age_max:
                # Calculate birth date for minimum age (older people)
                max_age = int(age_max)
                min_birth_date = date(today.year - max_age - 1, today.month, today.day)
                queryset = queryset.filter(date_of_birth__gte=min_birth_date)
            
            if age_min:
                # Calculate birth date for maximum age (younger people)
                min_age = int(age_min)
                max_birth_date = date(today.year - min_age, today.month, today.day)
                queryset = queryset.filter(date_of_birth__lte=max_birth_date)
        
        return queryset.order_by('last_name', 'first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get formal puroks from the Purok model
        from .models import Purok
        context['puroks'] = Purok.objects.values_list('name', flat=True).order_by('name')
        
        # Preserve filter values
        context['search_query'] = self.request.GET.get('search', '')
        context['purok_filter'] = self.request.GET.get('purok', '')
        context['sector_filter'] = self.request.GET.get('sector', '')
        context['age_min'] = self.request.GET.get('age_min', '')
        context['age_max'] = self.request.GET.get('age_max', '')
        
        # Total residents count
        context['total_residents'] = Resident.objects.filter(is_active=True).count()
        
        return context


@non_bootstrap_required
def export_residents_excel(request):
    """
    Export residents list to Excel file.
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Residents Masterlist"
    
    # Define headers
    headers = [
        'ID', 'Last Name', 'First Name', 'Middle Name', 'Suffix',
        'Date of Birth', 'Age', 'Sex', 'Civil Status',
        'Purok', 'Address', 'Mobile Number', 'Email',
        'Senior Citizen', 'PWD', 'Solo Parent', '4Ps', 'Voter',
        'Occupation', 'Educational Attainment'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get residents data
    residents = Resident.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    # Write data
    for row_num, resident in enumerate(residents, 2):
        ws.cell(row=row_num, column=1, value=resident.id)
        ws.cell(row=row_num, column=2, value=resident.last_name)
        ws.cell(row=row_num, column=3, value=resident.first_name)
        ws.cell(row=row_num, column=4, value=resident.middle_name)
        ws.cell(row=row_num, column=5, value=resident.suffix)
        ws.cell(row=row_num, column=6, value=resident.date_of_birth.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=resident.age)
        ws.cell(row=row_num, column=8, value=resident.get_sex_display())
        ws.cell(row=row_num, column=9, value=resident.get_civil_status_display())
        ws.cell(row=row_num, column=10, value=resident.purok)
        ws.cell(row=row_num, column=11, value=resident.address)
        ws.cell(row=row_num, column=12, value=resident.mobile_number)
        ws.cell(row=row_num, column=13, value=resident.email)
        ws.cell(row=row_num, column=14, value='Yes' if resident.is_senior_citizen else 'No')
        ws.cell(row=row_num, column=15, value='Yes' if resident.is_pwd else 'No')
        ws.cell(row=row_num, column=16, value='Yes' if resident.is_solo_parent else 'No')
        ws.cell(row=row_num, column=17, value='Yes' if resident.is_4ps else 'No')
        ws.cell(row=row_num, column=18, value='Yes' if resident.is_voter else 'No')
        ws.cell(row=row_num, column=19, value=resident.occupation)
        ws.cell(row=row_num, column=20, value=resident.get_educational_attainment_display() if resident.educational_attainment else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Residents_Masterlist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


class ResidentCreateView(LoginRequiredMixin, NonBootstrapRequiredMixin, CreateView):
    """
    View for creating new resident profiles.
    """
    model = Resident
    form_class = None  # Will import from forms
    template_name = 'pages/residents/form.html'
    
    def get_form_class(self):
        return ResidentForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Add'
        context['title'] = 'Add New Resident'
        return context
    
    def form_valid(self, form):
        
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Resident "{self.object.full_name}" has been successfully added!'
        )
        return response
    
    def get_success_url(self):
        return reverse('residents:list')


class ResidentUpdateView(LoginRequiredMixin, NonBootstrapRequiredMixin, UpdateView):
    """
    View for editing existing resident profiles.
    """
    model = Resident
    form_class = None  # Will import from forms
    template_name = 'pages/residents/form.html'
    
    def get_form_class(self):
        return ResidentForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Edit'
        context['title'] = f'Edit Resident: {self.object.full_name}'
        return context
    
    def form_valid(self, form):
        
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Resident "{self.object.full_name}" has been successfully updated!'
        )
        return response
    
    def get_success_url(self):
        return reverse('residents:list')


class ResidentDetailView(LoginRequiredMixin, NonBootstrapRequiredMixin, DetailView):
    """
    Detailed view of a resident's profile.
    """
    model = Resident
    template_name = 'pages/residents/detail.html'
    context_object_name = 'resident'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get household members if this person is a household head
        if self.object.is_household_head:
            context['household_members'] = Resident.objects.filter(
                household_head=self.object,
                is_active=True
            ).order_by('date_of_birth')
        
        # Get household head info if this person is a member
        if not self.object.is_household_head and self.object.household_head:
            context['household_head'] = self.object.household_head
        
        return context
class ResidentSearchView(LoginRequiredMixin, NonBootstrapRequiredMixin, ListView):
    model = Resident
    template_name = 'pages/residents/partials/search_results.html'
    context_object_name = 'residents'
    
    def get_template_names(self):
        if self.request.GET.get('source') == 'global':
            return ['partials/global_search_results.html']
        return [self.template_name]

    def get_queryset(self):
        query = self.request.GET.get('q', '').strip()
        if len(query) < 2:
            return Resident.objects.none()
        
        # Split query into words for individual field matching
        words = query.split()
        
        filters = Q(first_name__icontains=query) | \
                  Q(last_name__icontains=query) | \
                  Q(middle_name__icontains=query) | \
                  Q(id__icontains=query)
        
        # Add support for "First Last" search by matching multiple fields
        if len(words) >= 2:
            combined_filter = Q()
            for word in words:
                combined_filter &= (Q(first_name__icontains=word) | 
                                   Q(last_name__icontains=word) | 
                                   Q(middle_name__icontains=word))
            filters |= combined_filter

        return Resident.objects.filter(filters, is_active=True).order_by('last_name', 'first_name')[:10]
from django.shortcuts import render, redirect
from .forms import HouseholdHeadForm, HouseholdMemberFormSet

class HouseholdRegistrationView(LoginRequiredMixin, NonBootstrapRequiredMixin, CreateView):
    """
    View for bulk registration of a household (head + members).
    """
    model = Resident
    template_name = 'pages/residents/household_form.html'
    
    def get(self, request, *args, **kwargs):
        self.object = None
        form = HouseholdHeadForm()
        formset = HouseholdMemberFormSet()
        return self.render_to_response(
            self.get_context_data(form=form, formset=formset)
        )
    
    def post(self, request, *args, **kwargs):
        self.object = None
        form = HouseholdHeadForm(request.POST, request.FILES)
        formset = HouseholdMemberFormSet(request.POST, request.FILES)
        
        if form.is_valid() and formset.is_valid():
            return self.form_valid(form, formset)
        else:
            return self.form_invalid(form, formset)
    
    def form_valid(self, form, formset):
        # Save the household head first
        self.object = form.save()
        
        # Assign the head to the members and save them
        instances = formset.save(commit=False)
        for instance in instances:
            instance.household_head = self.object
            # Inherit address and purok from head if blank
            if not instance.purok:
                instance.purok = self.object.purok
            if not instance.address:
                instance.address = self.object.address
            instance.save()
        
        formset.save_m2m() # In case there are many-to-many fields
        
        messages.success(
            self.request,
            f'Household for "{self.object.full_name}" has been successfully registered!'
        )
        return redirect(self.get_success_url())
    
    def form_invalid(self, form, formset):
        return self.render_to_response(
            self.get_context_data(form=form, formset=formset)
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bulk Household Registration'
        return context
    
    def get_success_url(self):
        return reverse('residents:list')
